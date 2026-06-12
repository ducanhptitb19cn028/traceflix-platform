#!/usr/bin/env python3
"""Export the raw observability (MELT) data behind the RQ4 comparison to Excel.

Regenerates the same non-stationary telemetry stream used by
ml.experiments.online_vs_offline (same episodes/seed) and writes:

  MELT_Windows    one row per (service, time-window): every metric/log/trace/
                  event field, plus the ground-truth label and operational regime
  Features_C4     the engineered feature matrix the models actually consume (C4)
  Regime_Legend   regime index -> name / meaning

Also drops a flat CSV (observability_melt.csv) alongside the workbook.

Run:  python -m ml.eval.export_observability --episodes 320 --out data/results
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from ..drift import REGIME_NAMES, generate_drifting_run
from ..configs import CONFIGS
from ..features.build import build_features, METRIC_FEATS, LOG_FEATS, TRACE_FEATS, EVENT_FEATS


def _style(ws, ncols: int):
    fill = PatternFill("solid", fgColor="305496")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = fill, font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col[:200] if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 32)
    ws.freeze_panes = "A2"


def windows_to_df(windows, regimes) -> pd.DataFrame:
    rows = []
    for w, rg in zip(windows, regimes):
        row = {
            "ts": w.ts, "service": w.service,
            "label_fault": w.fault, "is_anomaly": int(w.fault != "normal"),
            "regime": rg, "regime_name": REGIME_NAMES[rg],
        }
        for pillar, names in (("metrics", METRIC_FEATS), ("logs", LOG_FEATS),
                              ("traces", TRACE_FEATS), ("events", EVENT_FEATS)):
            bag = getattr(w, pillar)
            for n in names:
                row[f"{pillar}.{n}"] = float(bag.get(n, 0.0))
        rows.append(row)
    return pd.DataFrame(rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--episodes", type=int, default=320)
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--out", default="data/results")
    args = ap.parse_args()
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    print(f"[*] Regenerating observability stream ({args.episodes} episodes, seed {args.seed})")
    windows, regimes = generate_drifting_run(n_episodes=args.episodes, seed=args.seed)
    melt = windows_to_df(windows, regimes)
    print(f"    {len(melt)} windows x {melt.shape[1]} columns")

    feats_c4 = build_features(windows, CONFIGS["C4"])

    legend = pd.DataFrame({
        "regime": list(range(len(REGIME_NAMES))),
        "regime_name": REGIME_NAMES,
        "meaning": [
            "baseline operating point (offline model trained here)",
            "release regresses latency; CPU and GC climb",
            "scale-out: throughput and log/trace volume double, memory grows",
            "combined heavy load: everything has shifted",
        ],
    })

    # flat CSV too
    melt.to_csv(out / "observability_melt.csv", index=False)

    xlsx = out / "observability_data.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xl:
        melt.round(4).to_excel(xl, sheet_name="MELT_Windows", index=False)
        feats_c4.round(4).to_excel(xl, sheet_name="Features_C4", index=False)
        legend.to_excel(xl, sheet_name="Regime_Legend", index=False)
        _style(xl.book["MELT_Windows"], melt.shape[1])
        _style(xl.book["Features_C4"], feats_c4.shape[1])
        _style(xl.book["Regime_Legend"], legend.shape[1])

    print(f"[*] Excel  -> {xlsx}")
    print(f"[*] CSV    -> {out / 'observability_melt.csv'}")
    print("\nLabel / regime distribution:")
    print(pd.crosstab(melt.regime_name, melt.is_anomaly).rename(
        columns={0: "normal", 1: "anomaly"}).to_string())


if __name__ == "__main__":
    main()
