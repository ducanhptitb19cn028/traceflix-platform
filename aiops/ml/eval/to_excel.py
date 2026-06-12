#!/usr/bin/env python3
"""Export the RQ4 offline-vs-online comparison to a formatted Excel workbook.

Reads the CSVs produced by ml.experiments.online_vs_offline and
ml.experiments.cost_compare, and writes a multi-sheet .xlsx:

  Summary        headline F1 (offline_static/periodic vs online_adaptive vs oracle)
  Detection_F1   config x model pivot of F1 on the operational future
  Detection_All  full per-(config, segment, model) metrics table
  PerRegime_F1   F1 broken down by future regime R1..R3
  Cost           latency / model-size / retained-data trade-off
  Timeline       block-wise rolling F1 over the drifting stream

Run:  python -m ml.eval.to_excel data/results
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MODEL_ORDER = ["offline_static", "offline_periodic", "online_adaptive", "offline_full"]
MODEL_LABEL = {
    "offline_static": "Offline (static)",
    "offline_periodic": "Offline (periodic retrain)",
    "online_adaptive": "Online (adaptive)",
    "offline_full": "Offline (oracle ceiling)",
}


def _autofit_and_style(ws, df_ncols: int, freeze: str = "A2"):
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    for c in range(1, df_ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # column widths from max content length
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 40)
    ws.freeze_panes = freeze


def main():
    results = Path(sys.argv[1] if len(sys.argv) > 1 else "data/results")
    det = pd.read_csv(results / "rq4_online_vs_offline.csv")
    cost = pd.read_csv(results / "rq4_cost.csv")
    timeline = pd.read_csv(results / "rq4_timeline.csv")

    # --- headline: F1 on the operational future, config x model ---
    future = det[det.segment.isin(["overall_future", "overall_allregimes"])].copy()
    future["model"] = pd.Categorical(future["model"], MODEL_ORDER, ordered=True)
    f1_pivot = (future.pivot_table(index=["config", "name"], columns="model",
                                   values="f1", observed=True)
                      .reindex(columns=MODEL_ORDER).reset_index())
    f1_pivot["online_gain_vs_static"] = (
        f1_pivot["online_adaptive"] - f1_pivot["offline_static"])
    f1_pivot["online_gain_vs_periodic"] = (
        f1_pivot["online_adaptive"] - f1_pivot["offline_periodic"])

    # --- summary (long, friendly labels + all four metrics on the future) ---
    summary = future[["config", "name", "model", "precision", "recall", "f1",
                      "auc_roc", "n"]].copy()
    summary["model"] = summary["model"].map(MODEL_LABEL)
    summary = summary.sort_values(["config", "f1"]).reset_index(drop=True)

    # --- per-regime F1 (R1..R3) ---
    regime = det[det.regime >= 0].copy()
    regime["model"] = pd.Categorical(regime["model"], MODEL_ORDER, ordered=True)
    perreg = (regime.pivot_table(index=["config", "segment"], columns="model",
                                 values="f1", observed=True)
                    .reindex(columns=[m for m in MODEL_ORDER if m != "offline_full"])
                    .reset_index())

    out = results / "rq4_offline_vs_online_comparison.xlsx"
    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        f1_pivot.round(4).to_excel(xl, sheet_name="Summary", index=False)
        f1_pivot.round(4).to_excel(xl, sheet_name="Detection_F1", index=False)
        summary.round(4).to_excel(xl, sheet_name="Detection_All", index=False)
        perreg.round(4).to_excel(xl, sheet_name="PerRegime_F1", index=False)
        cost.round(4).to_excel(xl, sheet_name="Cost", index=False)
        timeline.round(4).to_excel(xl, sheet_name="Timeline", index=False)

        for name, ncols in [("Summary", f1_pivot.shape[1]),
                            ("Detection_F1", f1_pivot.shape[1]),
                            ("Detection_All", summary.shape[1]),
                            ("PerRegime_F1", perreg.shape[1]),
                            ("Cost", cost.shape[1]),
                            ("Timeline", timeline.shape[1])]:
            _autofit_and_style(xl.book[name], ncols)

    print(f"[*] Excel workbook -> {out}")
    print("\nHeadline F1 (operational future R1-R3):")
    print(f1_pivot.round(3).to_string(index=False))


if __name__ == "__main__":
    main()
